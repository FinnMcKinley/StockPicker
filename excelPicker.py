import random
from datetime import date
from openpyxl import load_workbook
from openpyxl import Workbook

today = date.today()
dateToday = today.strftime("%d/%m/%y")
randomNumber = random.randint(2,749)
print(randomNumber)
print(dateToday)
randomCell = 'H' + str(randomNumber)
print('randomCell is:')
print(randomCell)

InfoWB = load_workbook(filename= 'ExcelSheet.xlsx')
InfoWS = InfoWB.active
randomStock = InfoWS[randomCell].value
InfoWB.save('ExcelSheet.xlsx')
InfoWB.close()
print(randomStock)

dbWB = load_workbook(filename= 'StockPicks.xlsx')
db = dbWB.active
# This next bit doesn't work either but I haven't researched or tried fixing it yet
# dbWB['A2'] = randomStock
# dbWB.__setattr__('A2', randomStock)
# db['A2'] = randomStock
# db.cell(row=2, column=1, value=randomStock)
dbWB.save('StockPicks.xlsx')
dbWB.close()
